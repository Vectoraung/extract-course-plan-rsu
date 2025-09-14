from extract_from_gpa_file.models import CourseGroup, Course
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font, Alignment

import openpyxl

from extract_from_gpa_file.exact_subjects import major_subjects, core_subjects, major_elective_subjects, rsu_identity_subjects
from extract_from_gpa_file.scrape_subjects import start_scrapping, get_all_subjects
from extract_from_gpa_file.excel_configs import excel_configs

final_tables = start_scrapping("Rangsit University.pdf")
final_tables = final_tables[1:]
all_subjects = get_all_subjects(final_tables)

major_courses = CourseGroup(
    name="Major Courses",
    total_credits=60,
    allowed_courses=major_subjects,
    )

core_courses = CourseGroup(
    name="Core Courses",
    total_credits=9,
    allowed_courses=core_subjects
    )

major_elective = CourseGroup(
    name="Major Electives",
    total_credits=15,
    allowed_courses=major_elective_subjects
    )

rsu_identity = CourseGroup(
    name="RSU Identity",
    total_credits=3,
    allowed_courses=rsu_identity_subjects
    )

general_education = CourseGroup(
    name="General Education",
    total_credits=15,
    allowed_group_numbers=[3, 4, 5, 6, 7, 8]
    )

ic = CourseGroup(
    name="Internationalization and Communication",
    total_credits=12,
    allowed_group_numbers=[1, 2]
    )

for subj in all_subjects:
    course = Course(
        subj["code"],
        subj["credit"],
        subj["grade"],
        subj["semester"],
        subj["year_eng"],
        subj["year_thai"]
    )
    
    success, m = major_courses.add_course(course)
    print(m)
    if success:
        print("added to major ", course.code)
        continue

    success, m = core_courses.add_course(course)
    print(m)
    if success:
        print("added to core ", course.code)
        continue

    success, m = major_elective.add_course(course)
    print(m)
    if success:
        print("added to elective ", course.code)
        continue

    success, m = rsu_identity.add_course(course)
    print(m)
    if success:
        print("added to rsu identity ", course.code)
        continue

    success, m = general_education.add_course(course)
    print(m)
    if success:
        print("added to general education ", course.code)
        continue

    success, m = ic.add_course(course)
    print(m)
    if success:
        print("added to ic ", course.code)
        continue

wb = openpyxl.Workbook()
ws = wb.active

all_courses = [rsu_identity, ic, general_education, core_courses, major_courses, major_elective]
colors = [
    'FFC4B3',  # Soft Coral
    'B8E2B8',  # Soft Green
    'D9D9D9',  # Soft Gray
    'A8DADC',  # Soft Cyan
    'FFE8A1',  # Soft Yellow
    'F5C1E0'   # Soft Pink
]

start_row = 1
for i, course_group in enumerate(all_courses):
    #print("start ", start_row)
    course_group.show_courses()
    show_all_allowed_courses=False
    if course_group.name == "Major Courses":
        show_all_allowed_courses=True
    excel_data = course_group.convert_to_excel_format(show_all_allowed_courses)

    # Append data
    for row in excel_data:
        ws.append(row)

    ws.append([])

    end_row = start_row + len(excel_data)
    #print("end ", end_row)

    # Make header row bold
    bold_font = Font(bold=True)
    ws.cell(row=start_row, column=1).font = bold_font

    # Fill color to cells from start row to end row
    fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type='solid')
    border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    for col in range(1, ws.max_column + 1):
        if i == 1:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = excel_configs[col]["width"]
        for row in range(start_row, end_row):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # Merge cells for header
    ws.merge_cells(f"{openpyxl.utils.get_column_letter(1)}{start_row}:{openpyxl.utils.get_column_letter(ws.max_column-2)}{start_row}")
    ws.merge_cells(f"{openpyxl.utils.get_column_letter(7)}{start_row}:{openpyxl.utils.get_column_letter(ws.max_column)}{start_row}")

    start_row = end_row + 1

wb.save("courses.xlsx")