import openpyxl
import json
import io

class CoursePlanFitter:
    def __init__(self):
        self.student_name = None
        self.student_id = None
        self.faculty_name = None
        self.information = None

        self.courses = []

    def fit(self, data):
        self.information = data['information']

        self.student_name = data['information']['name']
        self.student_id = data['information']['student_id']
        self.faculty_name = data['information']['faculty']['name']

        for line in data['courses']:
            self.courses.append(self._format_course(line, term_full_word=True))

        ##################################################################################
        dummy_courses = ["ICT212", "ICT210", "ICT422", "ICT336"]

        for course in self.courses:
            for c in dummy_courses:
                if course['code'] == c:
                    index = self.courses.index(course)
                    self.courses[index]['grade'] = "A"
                    self.courses[index]['credit'] = 3
                    self.courses[index]['term_number'] = 1
                    self.courses[index]['year_eng'] = 2025
                    self.courses[index]['year_thai'] = 2568
                    self.courses[index]['term'] = "first / 2025"
        ###################################################################################

        self.courses.sort(
            key=lambda c: (
                c["year_eng"] is None,   # False (0) first, True (1) last
                c["year_eng"] if c["year_eng"] is not None else float("inf"),
                c["term_number"] is None,       # False (0) first, True (1) last
                c["term_number"] if c["term"] is not None else float("inf")
            )
        )

    def _format_course(self, line, term_full_word=False):
        course = {
            "code": None,
            "credit": None,
            "grade": None,
            "term_number": None,
            "year_eng": None,
            "year_thai": None,
            "term": None
        }

        line = line.split(" ")
        course['code'] = line[1]
        course['credit'] = int(line[2])

        if len(line) < 4:
            return course

        course['grade'] = line[3]

        sem_and_year = line[4].split("/")

        course['term_number'], course['year_thai'] = int(sem_and_year[0]), int(sem_and_year[1])

        course['year_eng'] = self._thai_year_to_english(thai_year=course['year_thai'])

        term = None
        if course['term_number'] == 1:
            term = "first"
        elif course['term_number'] == 2:
            term = "second"
        elif course['term_number'] == 3:
            term = "summer"
        else:
            term = "UNKNOWN"

        if term_full_word:
            course['term'] = f"{term} / {course['year_eng']}"
            return course

        course['term'] = f"{'S' if course['term_number'] == 3 else course['term_number']} / {course['year_eng']}"

        return course

    def _thai_year_to_english(self, thai_year, month=1):
        """
        Convert Thai Buddhist Era year to Gregorian year.
        """
        if month <= 3:
            return thai_year - 544
        else:
            return thai_year - 543

    def _get_course_plan_frame_name(self):
        try:
            path = 'course_plan_excel_frame/'
            frame_name = path + self.faculty_name + '.xlsx'

            wb = openpyxl.load_workbook(frame_name)
            return wb
        except:
            return None
        
    def _append_ge_courses(self, wb, col_configs):
        if wb is None:
            return None
        
        ws = wb.active

        with open("database/ge_excel_configs.json", "r", encoding="utf-8") as f:
            ge_configs = json.load(f)

        group_3_to_8_titles = [
            "Leadership and Social Responsibility",
            "Arts and Culture",
            "Innovative Entrepreneurship",
            "Digital Media Literacy",
            "Essence of Science",
            "RSU My-Style"
        ]

        free_electives = []
        overlapped_courses = []

        added_courses = {}
        left_courses = []

        ### Adding GE Courses ###
        for course in self.courses:
            group_name = self._fine_ge_group(course['code'], ge_configs)

            if group_name is None:
                left_courses.append(course)
                continue      

            counter = len(set(added_courses.keys()) & set(group_3_to_8_titles)) 

            if counter >= 5:
                free_electives.append(course)
                continue

            row = 0
            if group_name not in added_courses:
                added_courses[group_name] = []

            row = ge_configs[group_name]["start_row"] + len(added_courses[group_name])

            if course in added_courses[group_name]:
                #if group_name not in overlapped_courses:
                    #overlapped_courses[group_name] = []
                #overlapped_courses[group_name].append(course)
                overlapped_courses.append(course)
                continue

            total_credits = 0
            for added_course in added_courses[group_name]:
                total_credits += added_course['credit']

            if total_credits >= ge_configs[group_name]["max_credit"]:
                #if group_name not in free_electives:
                    #free_electives[group_name] = []
                #free_electives[group_name].append(course)
                free_electives.append(course)
                continue            

            if group_name in group_3_to_8_titles and group_name in added_courses and len(added_courses[group_name]) == 1:
                overlapped_courses.append(course)
                continue

            added_courses[group_name].append(course)

            for col_name, config in col_configs.items():
                if col_name == 'number':
                    continue
                ws.cell(row=row, column=config["col"]).value = course[col_name]

        return wb, added_courses, left_courses, free_electives, overlapped_courses

    def _append_specialized_major_courses(self, wb, col_configs):
        if wb is None:
            return None
        
        ws = wb.active

        with open("database/specialized_major_excel_configs.json", "r", encoding="utf-8") as f:
            specialized_majors = json.load(f)

        group_excel_configs = specialized_majors.get(self.faculty_name, None)
        if group_excel_configs is None:
            return {}, [], [], [], []

        overlapped_courses = []
        free_electives = []

        added_courses = {}
        left_courses = []

        ### Adding Specialized Major Courses ###
        for course in self.courses:
            group_name = None

            for name, rule in group_excel_configs.items():
                if course['code'] in rule["courses"]:
                    group_name = name
                    break

            if group_name is None:
                left_courses.append(course)
                continue

            if not group_excel_configs[group_name]['show_all'] and course['grade'] is None:
                continue

            row = 0
            if group_name not in added_courses:
                added_courses[group_name] = []

            if group_excel_configs[group_name]['show_all']:
                row = group_excel_configs[group_name]["start_row"] + group_excel_configs[group_name]["courses"].index(course['code'])
            else:
                row = group_excel_configs[group_name]["start_row"] + len(added_courses[group_name])

            if course in added_courses[group_name]:
                #if group_name not in overlapped_courses:
                    #overlapped_courses[group_name] = []
                #overlapped_courses[group_name].append(course)
                overlapped_courses.append(course)
                continue

            total_credits = 0
            for added_course in added_courses[group_name]:
                total_credits += added_course['credit']

            if total_credits >= group_excel_configs[group_name]["max_credit"]:
                #if group_name not in free_electives:
                    #free_electives[group_name] = []
                #free_electives[group_name].append(course)
                free_electives.append(course)
                continue

            added_courses[group_name].append(course)

            for col_name, config in col_configs.items():
                if col_name == 'number':
                    continue
                ws.cell(row=row, column=config["col"]).value = course[col_name]

        return wb, added_courses, left_courses, free_electives, overlapped_courses

    def generate_excel_file(self, is_web=False):
        wb = self._get_course_plan_frame_name()

        with open("database/excel_column_configs.json", "r", encoding="utf-8") as f:
            col_configs = json.load(f)

        wb, added_ge_courses, left_ge_courses, free_electives_ge, ol_ge_courses = self._append_ge_courses(wb, col_configs)
        wb, added_specialized_courses, left_specialized_courses, free_electives_s, ol_s_courses = self._append_specialized_major_courses(wb, col_configs)

        set1 = {frozenset(d.items()) for d in left_ge_courses}
        set2 = {frozenset(d.items()) for d in left_specialized_courses}

        # Intersection
        common = set1 & set2
        all_left_courses = [dict(items) for items in common]
        all_free_electives = free_electives_ge + free_electives_s + all_left_courses

        all_free_electives.sort(
            key=lambda c: (
                c["year_eng"] is None,   # False (0) first, True (1) last
                c["year_eng"] if c["year_eng"] is not None else float("inf"),
                c["term_number"] is None,       # False (0) first, True (1) last
                c["term_number"] if c["term"] is not None else float("inf")
            )
        )

        with open("database/free_electives_and_false_courses.json", "r", encoding="utf-8") as f:
            configs = json.load(f)

        ws = wb.active

        free_electives_total_credits = 0
        added_false_courses = []

        for course in all_free_electives:
            row = 0
            if not free_electives_total_credits >= configs["Free Electives"]["max_credit"]:
                row = configs['Free Electives']['start_row'] + all_free_electives.index(course)
                free_electives_total_credits += course['credit']
            else:
                row = configs['False Courses']['start_row'] + len(added_false_courses)
                print("row ", row)
                print(len(added_false_courses))
                added_false_courses.append(course)       

            for col_name, config in col_configs.items():
                if col_name == 'number':
                    continue
                ws.cell(row=row, column=config["col"]).value = course[col_name]

        all_ol_courses = ol_ge_courses + ol_s_courses

        for course in all_ol_courses:
            row = configs['False Courses']['start_row'] + len(added_false_courses)

            for col_name, config in col_configs.items():
                if col_name == 'number':
                    continue
                ws.cell(row=row, column=config["col"]).value = course[col_name]

            added_false_courses.append(course)

        file_name = f"{self.student_name}_{self.student_id}_course_plan"

        if is_web:
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)  # rewind to the start

            return excel_buffer, self.information, added_ge_courses, added_specialized_courses
        
        wb.save(f"{file_name}.xlsx")
    
    def _fine_ge_group(self, course_code, rules):
        # Extract prefix (letters) and number part (digits)
        prefix = ''.join([c for c in course_code if c.isalpha()])
        digits = ''.join([c for c in course_code if c.isdigit()])

        # Ensure digits are available
        if len(digits) < 2:
            return None

        second_digit = digits[1]  # second digit of the number part

        # Check each group
        for group_name, rule in rules.items():
            if prefix in rule["prefix"] and second_digit == rule["second_digit"]:
                return group_name

        return None

        

