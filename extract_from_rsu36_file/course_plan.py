import pdfplumber
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font, Alignment
import openpyxl
import io
import re

from extract_from_rsu36_file.course import Course
from extract_from_rsu36_file.course_group import CourseGroup

excel_configs = {
    1: {"width": 5},
    2: {"width": 15},
    3: {"width": 15},
    4: {"width": 10},
    5: {"width": 10},
    6: {"width": 20},
    7: {"width": 10},
    8: {"width": 10},
}
    
class CoursePlan:
    def __init__(self):
        self.groups = []

        self.student_name = None
        self.student_id = None

        self.major = self._create_major_group("Major Courses", 60)
        self.core = self._create_core_group("Core Courses", 9)
        self.major_elective = self._create_major_elective_group("Major Electives", 15)
        self.rsu_i = self._create_rsu_i_group("RSU Identity", 3)
        self.ge = self._create_ge_group("GE Courses", 15)
        self.ic = self._create_ic_group("IC Courses", 12)

    def import_with_RSU36_file(self, file_obj):
        pdf = pdfplumber.open(file_obj)
        page = pdf.pages[0]

        all_text = self._split_half(page)
        prefixes_to_remove = [
            "1หมวดววชาศศกษาทลสวไป",
            "2หมวดววชาเฉพาะ",
        ]

        all_text = [
            t for t in self._split_half(page)
            if not any(t.startswith(prefix) for prefix in prefixes_to_remove)
        ]

        groups = self._extract_courses(all_text)
        
        self._extract_student_name(all_text)
        self._extract_student_id(all_text)

        except_ge_group_names = [
            "อลตลลกษณณมหาววทยาลลย",
            "ความณเปนสากลและการสสสอสาร",
            "กณลมล ววชาพสพนฐานววชาชยพ",
            "กณลมล ววชาชยพบลงคลบ",
            "กณลมล ววชาชยพเลสอก",
        ]
        '''except_ge_group_names = [
            "อลตลลกษณณมหาววทยาลลย 3 /3 หนนวยกกต",
            "ความณเปนสากลและการสสสอสาร 12 /12 หนนวยกกต",
            "2.1 กณลมล ววชาพสพนฐานววชาชยพ 9 /9 หนนวยกกต",
            "2.2 กณลมล ววชาชยพบลงคลบ 48 /60 หนนวยกกต",
            "2.3 กณลมล ววชาชยพเลสอก 9 /15 หนนวยกกต",
        ]'''

        for course in groups["อลตลลกษณณมหาววทยาลลย"]:
            self.rsu_i.add_course(course=Course(raw_line=course))

        for course in groups["ความณเปนสากลและการสสสอสาร"]:
            self.ic.add_course(Course(raw_line=course))

        for course in groups["กณลมล ววชาพสพนฐานววชาชยพ"]:
            self.core.add_course(Course(raw_line=course))

        for course in groups["กณลมล ววชาชยพบลงคลบ"]:
            self.major.add_course(Course(raw_line=course))

        for course in groups["กณลมล ววชาชยพเลสอก"]:
            self.major_elective.add_course(Course(raw_line=course))

        for g in groups:
            if g not in except_ge_group_names:
                for course in groups[g]:
                    self.ge.add_course(Course(raw_line=course))

    def generate_excel_file(self, streamlit=False):
        wb = openpyxl.Workbook()
        ws = wb.active

        configs = {
            "rsu_i": {"obj": self.rsu_i, "color": "A8DADC", "show_only_finished_courses": False},
            "ic": {"obj": self.ic, "color": "F5C1E0", "show_only_finished_courses": False},
            "ge": {"obj": self.ge, "color": "FFE8A1", "show_only_finished_courses": False},
            "core": {"obj": self.core, "color": "B8E2B8", "show_only_finished_courses": False},
            "major": {"obj": self.major, "color": "FFC4B3", "show_only_finished_courses": False},
            "major_elective": {"obj": self.major_elective, "color": "D9D9D9", "show_only_finished_courses": True},
        }

        start_row = 1
        for name, config in configs.items():
            excel_data = config["obj"].convert_to_excel_format(show_only_finished_courses=config["show_only_finished_courses"])
            for row in excel_data:
                ws.append(row)
            end_row = start_row + len(excel_data)

            bold_font = Font(bold=True)
            ws.cell(row=start_row, column=1).font = bold_font

            fill = PatternFill(start_color=config["color"], end_color=config["color"], fill_type='solid')
            border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))

            for col in range(1, ws.max_column + 1):
                if name == "rsu_i":
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = excel_configs[col]["width"]
                for row in range(start_row, end_row):
                    ws.cell(row=row, column=col).fill = fill
                    ws.cell(row=row, column=col).border = border
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

            # Merge cells for header
            ws.merge_cells(f"{openpyxl.utils.get_column_letter(1)}{start_row}:{openpyxl.utils.get_column_letter(ws.max_column-2)}{start_row}")
            ws.merge_cells(f"{openpyxl.utils.get_column_letter(7)}{start_row}:{openpyxl.utils.get_column_letter(ws.max_column)}{start_row}")

            ws.append([])

            start_row = end_row + 1

        file_name = f"{self.student_name}_{self.student_id}_courses_plan"

        if streamlit:
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)  # rewind to the start

            return excel_buffer, file_name
        wb.save(f"{file_name}.xlsx")

    def _extract_student_name(self, lines):
        name_line = lines[2]
        match = re.search(r"(MISS|MR\.)", name_line)
        if match:
            start_index = match.end()
            # Extract the name from the start index to the end of the string
            name = name_line[start_index:].strip().replace(" ", "_")
            self.student_name = name

    def _extract_student_id(self, lines):
        for line in lines:
            if line.startswith("รหหสประจจาตหว"):
                self.student_id = line.split(" ")[1]

    def _clean_group_name(self, text):
        string_to_remove = ["หนนวยกกต", ".", "/"]

        for s in string_to_remove:
            text = text.replace(s, "")

        text = re.sub(r'\d+', ' ', text).strip()

        return text

    def _extract_courses(self, lines):
        groups = {
        }
        current_group = None
        for i, line in enumerate(lines):
            if str(line).strip() == "รหหสวกชา หนนวยกกต เกรด ยชนยหน":
                temp = ""
                back_step = 1
                while not str(lines[i-back_step]).strip().startswith("ชชชอ") and not self._is_course_format(lines[i-back_step]):
                    temp = lines[i-back_step] + " " + temp
                    back_step += 1

                current_group = self._clean_group_name(temp)

                groups[current_group] = []

            if current_group is not None and str(line).strip() != "รหหสวกชา หนนวยกกต เกรด ยชนยหน" and self._is_course_format(line):
                groups[current_group].append(line)

        return groups

    def _is_course_format(self, line):
        parts = line.split(" ")
        if len(parts) < 3:
            return False
        if not parts[0].isnumeric():
            return False
        if not parts[1].isalnum():
            return False
        if not parts[2].isdigit() or not 1 <= int(parts[2]) <= 6:
            return False
        
        return True

    def _split_half(self, page):
        width = page.width
        height = page.height

        left_bbox = (0, 0, width / 2, height)
        left_crop = page.within_bbox(left_bbox)
        left_text = left_crop.extract_text()
        left_text_list = left_text.split("\n")
        #print("===== Left Half =====")
        #print(left_text)

        # Right half bounding box: (x0, y0, x1, y1)
        right_bbox = (width / 2, 0, width, height)
        right_crop = page.within_bbox(right_bbox)
        right_text = right_crop.extract_text()
        right_text_list = right_text.split("\n")
        #print("===== Right Half =====")
        #print(right_text)

        return left_text_list + right_text_list

    def _create_major_group(self, name, total_credits):
        group = CourseGroup(name, total_credits)
        self.groups.append(group)
        return group
    
    def _create_core_group(self, name, total_credits):
        group = CourseGroup(name, total_credits)
        self.groups.append(group)
        return group
    
    def _create_major_elective_group(self, name, total_credits):
        group = CourseGroup(name, total_credits)
        self.groups.append(group)
        return group
    
    def _create_rsu_i_group(self, name, total_credits):
        group = CourseGroup(name, total_credits)
        self.groups.append(group)
        return group
    
    def _create_ge_group(self, name, total_credits):
        group = CourseGroup(name, total_credits)
        self.groups.append(group)
        return group
    
    def _create_ic_group(self, name, total_credits):
        group = CourseGroup(name, total_credits)
        self.groups.append(group)
        return group
    

# EXAMPLE USAGE
#cp = CoursePlan()
#cp.import_with_RSU36_file("RSU36_form.pdf")

#cp.generate_excel_file()