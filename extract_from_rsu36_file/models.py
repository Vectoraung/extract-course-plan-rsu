import pdfplumber
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font, Alignment
import openpyxl
import io

excel_configs = {
    1: {"width": 5},
    2: {"width": 15},
    3: {"width": 15},
    4: {"width": 10},
    5: {"width": 10},
    6: {"width": 20},
    7: {"width": 10},
    8: {"width": 10},
}

class Course:
    def __init__(self, raw_line):
        self.code = None
        self.credit = None
        self.grade = None
        self.term = None
        self.year_eng = None
        self.year_thai = None

        if raw_line:
            self._format_from_raw_line(raw_line)

    def _format_from_raw_line(self, line):
        line = line.split(" ")
        self.code = line[1]
        self.credit = int(line[2])

        if len(line) < 4:
            return

        self.grade = line[3]

        sem_and_year = line[4].split("/")

        semester_num, year_thai = int(sem_and_year[0]), int(sem_and_year[1])

        if semester_num == 1:
            self.term = "FIRST SEMESTER"
        elif semester_num == 2:
            self.term = "SECOND SEMESTER"
        elif semester_num == 3:
            self.term = "SUMMER SESSION"
        else: self.term = "unknown"

        self.year_eng = self._thai_year_to_english(thai_year=year_thai)
        self.year_thai = year_thai
    
    def _thai_year_to_english(self, thai_year, month = 1):
        """
        Convert Thai Buddhist Era year to Gregorian year.
        
        Args:
            thai_year (int): Thai year (e.g., 2568)
            month (int): Month number (default=1)
            
        Returns:
            int: English year
        """
        if month <= 3:
            return thai_year - 544
        else:
            return thai_year - 543

    def __repr__(self):
        return f"{self.code} ({self.credit} cr) - {self.grade} [{self.term}]"


class CourseGroup:
    def __init__(self, name, total_credits):
        self.name = name
        self.total_credits = total_credits
        self.courses = []

    def add_course(self, course):
        self.courses.append(course)

    def get_total_credits(self):
        """Sum of credits from all courses in this group"""
        total = 0
        for c in self.courses:
            if c.grade is not None:
                total += c.credit
        return total

    def get_courses_by_term(self, term):
        """Return courses taken in a specific term"""
        return [c for c in self.courses if c.term == term]
    
    def show_courses(self, show_only_finished_courses=False):
        print(f"\n📚 {self.name} (Max Credits: {self.total_credits})")
        print("-" * 60)
        if not self.courses:
            print("No courses added yet.")
            return

        # Sort by code (e.g. ICT101 < ICT205 < IRS111)
        sorted_courses = sorted(self.courses, key=lambda c: c.code)

        # Filter if only finished courses are requested
        if show_only_finished_courses:
            sorted_courses = [c for c in sorted_courses if c.grade is not None]

        if not sorted_courses:
            print("No finished courses yet." if show_only_finished_courses else "No courses added yet.")
            return

        for i, c in enumerate(sorted_courses, 1):
            print(f"{i}. {c.code:<8} | {c.credit} Credits | Grade: {c.grade}"
                f" | Term: {c.term} | Year(EN): {c.year_eng} | Year(TH): {c.year_thai}")

        print(f"--> Total Credits so far: {self.get_total_credits()}")

    def convert_to_excel_format(self, show_only_finished_courses=False):
        """
        Convert courses to a 2D list (Excel-friendly format) with styling:
        - Title row (group name + max credits)
        - Header row
        - Rows for each course
        - Footer row with total credits
        """
        courses = self.courses

        courses = self.courses

        if show_only_finished_courses:
            courses = [c for c in courses if c.grade is not None]

        data = []

        current_total_credits = self.get_total_credits()

        # Title row
        title_text = f"{self.name} (Max Credits: {self.total_credits})"
        note_text = ""
        if current_total_credits == self.total_credits:
            note_text += " (FINISHED!)"
        else:
            note_text += f" (LEFT CREDITS: {self.total_credits - current_total_credits})"
        title = [title_text, "", "", "", "", "", note_text, ""]
        data.append(title)

        # Empty row for spacing
        data.append([])

        # Header row
        header = ["#", "Course Code", "Course Name", "Credits", "Grade", "Term", "Year (EN)", "Year (TH)"]
        data.append(header)

        # Ensure sorted order
        sorted_courses = sorted(courses, key=lambda c: c.code)

        # Course rows
        for i, c in enumerate(sorted_courses, 1):
            row = [
                i,
                c.code,
                getattr(c, "name", ""),  # optional course name if available
                c.credit,
                "" if c.grade is None else c.grade,
                "" if c.term is None else c.term,
                "" if c.year_eng == 0 else c.year_eng,
                "" if c.year_thai == 0 else c.year_thai
            ]
            data.append(row)

        # Footer row: total credits
        total_row = ["", "", "Total Credits", current_total_credits]
        
        data.append(total_row)
        data.append([])

        return data

    def calculate_gpa(self):
        """Weighted GPA for this group"""
        grade_points = {
            "A": 4.0, "B+": 3.5, "B": 3.0,
            "C+": 2.5, "C": 2.0,
            "D+": 1.5, "D": 1.0, "F": 0.0
        }
        total_points = 0
        total_credits = 0
        for c in self.courses:
            if c.grade in grade_points:
                total_points += grade_points[c.grade] * c.credit
                total_credits += c.credit
        return round(total_points / total_credits, 2) if total_credits > 0 else None

    def __repr__(self):
        return f"{self.name} ({self.get_total_credits()} / {self.total_credits} credits)"
    
class CoursePlan:
    def __init__(self):
        self.groups = []

        self.student_name = None
        self.student_id = None

        self.major = self._create_major_group("Major Courses", 60)
        self.core = self._create_core_group("Core Courses", 9)
        self.major_elective = self._create_major_elective_group("Major Electives", 15)
        self.rsu_i = self._create_rsu_i_group("RSU Identity", 3)
        self.ge = self._create_ge_group("GE Courses", 15)
        self.ic = self._create_ic_group("IC Courses", 12)

    def import_with_RSU36_file(self, file_obj):
        pdf = pdfplumber.open(file_obj)
        page = pdf.pages[0]

        all_text = self._split_half(page)

        groups = self._extract_courses(all_text)

        except_ge_group_names = [
            "อลตลลกษณณมหาววทยาลลย 3 /3 หนนวยกกต",
            "ความณเปนสากลและการสสสอสาร 12 /12 หนนวยกกต",
            "2หมวดววชาเฉพาะ 66 / 84 หนนวยกกต 2.1 กณลมล ววชาพสพนฐานววชาชยพ 9 /9 หนนวยกกต",
            "2.2 กณลมล ววชาชยพบลงคลบ 48 /60 หนนวยกกต",
            "2.3 กณลมล ววชาชยพเลสอก 9 /15 หนนวยกกต",
        ]

        for course in groups["อลตลลกษณณมหาววทยาลลย 3 /3 หนนวยกกต"]:
            self.rsu_i.add_course(course=Course(raw_line=course))

        for course in groups["ความณเปนสากลและการสสสอสาร 12 /12 หนนวยกกต"]:
            self.ic.add_course(Course(raw_line=course))

        for course in groups["2หมวดววชาเฉพาะ 66 / 84 หนนวยกกต 2.1 กณลมล ววชาพสพนฐานววชาชยพ 9 /9 หนนวยกกต"]:
            self.core.add_course(Course(raw_line=course))

        for course in groups["2.2 กณลมล ววชาชยพบลงคลบ 48 /60 หนนวยกกต"]:
            self.major.add_course(Course(raw_line=course))

        for course in groups["2.3 กณลมล ววชาชยพเลสอก 9 /15 หนนวยกกต"]:
            self.major_elective.add_course(Course(raw_line=course))

        for g in groups:
            if g not in except_ge_group_names:
                for course in groups[g]:
                    self.ge.add_course(Course(raw_line=course))

    def generate_excel_file(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        configs = {
            "rsu_i": {"obj": self.rsu_i, "color": "A8DADC", "show_only_finished_courses": False},
            "ic": {"obj": self.ic, "color": "F5C1E0", "show_only_finished_courses": False},
            "ge": {"obj": self.ge, "color": "FFE8A1", "show_only_finished_courses": False},
            "core": {"obj": self.core, "color": "B8E2B8", "show_only_finished_courses": False},
            "major": {"obj": self.major, "color": "FFC4B3", "show_only_finished_courses": False},
            "major_elective": {"obj": self.major_elective, "color": "D9D9D9", "show_only_finished_courses": True},
        }

        start_row = 1
        for name, config in configs.items():
            excel_data = config["obj"].convert_to_excel_format(show_only_finished_courses=config["show_only_finished_courses"])
            for row in excel_data:
                ws.append(row)
            end_row = start_row + len(excel_data)

            bold_font = Font(bold=True)
            ws.cell(row=start_row, column=1).font = bold_font

            fill = PatternFill(start_color=config["color"], end_color=config["color"], fill_type='solid')
            border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))

            for col in range(1, ws.max_column + 1):
                if name == "rsu_i":
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = excel_configs[col]["width"]
                for row in range(start_row, end_row):
                    ws.cell(row=row, column=col).fill = fill
                    ws.cell(row=row, column=col).border = border
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

            # Merge cells for header
            ws.merge_cells(f"{openpyxl.utils.get_column_letter(1)}{start_row}:{openpyxl.utils.get_column_letter(ws.max_column-2)}{start_row}")
            ws.merge_cells(f"{openpyxl.utils.get_column_letter(7)}{start_row}:{openpyxl.utils.get_column_letter(ws.max_column)}{start_row}")

            ws.append([])

            start_row = end_row + 1

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)  # rewind to the start
        
        return excel_buffer
        #wb.save("courses.xlsx")

    def _extract_courses(self, lines):
        groups = {
        }
        current_group = None
        for i, line in enumerate(lines):
            if str(line).strip() == "รหหสวกชา หนนวยกกต เกรด ยชนยหน":
                temp = ""
                back_step = 1
                while str(lines[i-back_step]).strip() != "1หมวดววชาศศกษาทลสวไป 30 / 30 หนนวยกกต" and not self._is_course_format(lines[i-back_step]):
                    temp = lines[i-back_step] + " " + temp
                    back_step += 1

                current_group = temp.strip()

                groups[current_group] = []

            if current_group is not None and str(line).strip() != "รหหสวกชา หนนวยกกต เกรด ยชนยหน" and self._is_course_format(line):
                groups[current_group].append(line)

        return groups

    def _is_course_format(self, line):
        parts = line.split(" ")
        if len(parts) < 3:
            return False
        if not parts[0].isnumeric():
            return False
        if not parts[1].isalnum():
            return False
        if not parts[2].isdigit() or not 1 <= int(parts[2]) <= 6:
            return False
        
        return True

    def _split_half(self, page):
        width = page.width
        height = page.height

        left_bbox = (0, 0, width / 2, height)
        left_crop = page.within_bbox(left_bbox)
        left_text = left_crop.extract_text()
        left_text_list = left_text.split("\n")
        #print("===== Left Half =====")
        #print(left_text)

        # Right half bounding box: (x0, y0, x1, y1)
        right_bbox = (width / 2, 0, width, height)
        right_crop = page.within_bbox(right_bbox)
        right_text = right_crop.extract_text()
        right_text_list = right_text.split("\n")
        #print("===== Right Half =====")
        #print(right_text)

        return left_text_list + right_text_list

    def _create_major_group(self, name, total_credits):
        group = CourseGroup(name, total_credits)
        self.groups.append(group)
        return group
    
    def _create_core_group(self, name, total_credits):
        group = CourseGroup(name, total_credits)
        self.groups.append(group)
        return group
    
    def _create_major_elective_group(self, name, total_credits):
        group = CourseGroup(name, total_credits)
        self.groups.append(group)
        return group
    
    def _create_rsu_i_group(self, name, total_credits):
        group = CourseGroup(name, total_credits)
        self.groups.append(group)
        return group
    
    def _create_ge_group(self, name, total_credits):
        group = CourseGroup(name, total_credits)
        self.groups.append(group)
        return group
    
    def _create_ic_group(self, name, total_credits):
        group = CourseGroup(name, total_credits)
        self.groups.append(group)
        return group
    

# EXAMPLE USAGE
#cp = CoursePlan()
#cp.import_with_RSU36_file("RSU36_form.pdf")

#cp.generate_excel_file()